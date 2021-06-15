#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Python Used 3.9.0
from kivy.config import Config
Config.set('graphics', 'resizable', '0')
Config.set('kivy', 'exit_on_escape', '0')

import json
from datetime import datetime
from pathlib import Path
import kivy as kv
import openpyxl as opx
import matplotlib.pyplot as plt
from matplotlib import rc
import os
import seaborn

from PIL import Image
from numpy import arange
from kivy.clock import Clock
from kivy.app import App
from kivy.core.window import Window
from kivy.graphics import Color, Rectangle, Line
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button, ButtonBehavior
from kivy.uix.checkbox import CheckBox
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.stacklayout import StackLayout
from kivy.uix.textinput import TextInput
from kivy.utils import get_color_from_hex

kv.require('2.0.0')


class ModifiedTextInput(TextInput):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.multiline = False
        self.background_color = get_color_from_hex('#231123')
        self.font_size = 16
        self.foreground_color = get_color_from_hex('#FFFFFF')
        self.hint_text_color = get_color_from_hex('#D72638')


class AddButton(Button):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.text = '[b]ADD[/b]'
        self.markup = True
        self.background_normal = ''
        self.background_down = ''
        self.bg_color = get_color_from_hex('#D72638')
        self.background_color = self.bg_color
        self.font_size = 16
        self.color = get_color_from_hex('#231123')

    def on_release(self):
        self.background_color = self.bg_color


class bg_change(FloatLayout):
    def __init__(self, clr='#6E44FF', **kwargs):
        super().__init__(**kwargs)

        self.clr = clr
        with self.canvas.before:
            Color(rgb=get_color_from_hex(self.clr))
            self.rect = Rectangle(size=self.size, pos=self.pos)

        self.bind(size=self._update, pos=self._update)

    def _update(self, wid, val):
        self.rect.pos = wid.pos
        self.rect.size = wid.size


class EntryField(FloatLayout):
    now = datetime.now()
    which_day = now.strftime('%d')
    which_month = now.strftime('%m')
    which_year = now.strftime("%Y")
    year_path = os.path.join('storage', str(which_year))
    month_path = os.path.join(year_path, str(which_month))
    day_path = os.path.join(month_path, str(which_day))
    Path('storage').mkdir(exist_ok=True)
    Path(year_path).mkdir(exist_ok=True)
    Path(month_path).mkdir(exist_ok=True)
    wb_complete_path = f"{day_path}.xlsx"
    cols_list = ['Sale', 'Sale Time', 'Amount', 'Selling Price', 'Profit', 'Buying Price', 'Description']

    if not Path(wb_complete_path).is_file():
        new_wb = opx.Workbook()
        for i, v in enumerate(cols_list, 1):
            new_wb.active.cell(row=1, column=i, value=v)

        new_wb.save(wb_complete_path)
        new_wb.close()

    load_wb = opx.load_workbook(wb_complete_path)
    ws = load_wb.active

    def __init__(self, title, grp, categories=['Alfa', 'Touch'], **kwargs):
        super().__init__(**kwargs)

        self.grp = grp
        self.categories = categories
        self.title = title
        self.label_color = '#231123'  # as same as background color

        self.title_label = Label(
            text=f'[b]{self.title.upper()}[/b]',
            pos_hint={'x': 0.5, 'y': 0.94},
            font_size=20,
            markup=True,
            color=get_color_from_hex(self.label_color),
            size_hint=(0, 0))
        self.add_widget(self.title_label)

        self.tglbox1 = BoxLayout(orientation='horizontal',
                                 pos_hint={'x': 0.005, 'y': 0.83},
                                 size_hint=(0.6, 0.06))

        self.tglbtn1 = CheckBox(group=self.grp,
                                size_hint=(0.2, 1))
        self.tglbox1.add_widget(self.tglbtn1)

        self.tglbtn1_lbl = Label(text=self.categories[0],
                                 size_hint=(0.2, 1),
                                 font_size=17)
        self.tglbox1.add_widget(self.tglbtn1_lbl)
        self.add_widget(self.tglbox1)

        self.tglbox2 = BoxLayout(orientation='horizontal',
                                 pos_hint={'x': 0.005, 'y': 0.74},
                                 size_hint=(0.6, 0.06))

        self.tglbtn2 = CheckBox(group=self.grp,
                                size_hint=(0.2, 1))
        self.tglbox2.add_widget(self.tglbtn2)

        self.tglbtn2_lbl = Label(text=self.categories[1],
                                 size_hint=(0.2, 1),
                                 font_size=17)
        self.tglbox2.add_widget(self.tglbtn2_lbl)
        self.add_widget(self.tglbox2)

        self.amt_box = BoxLayout(orientation='vertical',
                                 size_hint=(1, 0.1),
                                 pos_hint={'x': 0.02, 'y': 0.51},)

        self.amt_label = Label(text='Amount:',
                               size_hint=(0.34, 0.25))
        self.amt_box.add_widget(self.amt_label)

        self.amt_input = ModifiedTextInput(size_hint=(0.34, 0.35),
                                           input_filter='int',
                                           text='1')
        self.amt_box.add_widget(self.amt_input)
        self.add_widget(self.amt_box)

        self.price_box = BoxLayout(orientation='vertical',
                                   size_hint=(0.9, 0.1),
                                   pos_hint={'x': 0.02, 'y': 0.31})
        
        self.price_label = Label(text='Price:',
                                 size_hint=(0.2, 0.25))
        self.price_box.add_widget(self.price_label)

        self.price = ModifiedTextInput(hint_text='10000',
                                       size_hint=(0.55, 0.35),
                                       input_filter='int')
        self.price_box.add_widget(self.price)
        self.add_widget(self.price_box)

        self.cardtypes = ['Small', 'Big', 'Emergency', 'Start']
        self.card_choose = Spinner(text=self.cardtypes[0],
                                   values=self.cardtypes,
                                   pos_hint={'x': 0.02, 'y': 0.58},
                                   size_hint=(0.95, 0.08),
                                   background_normal='',
                                   background_color=get_color_from_hex('#231123'),
                                   sync_height=True,
                                   text_autoupdate=True)

        self.bp_box = BoxLayout(orientation='vertical',
                                size_hint=(0.9, 0.1),
                                pos_hint={'x': 0.02, 'y': 0.18})

        self.bp_label = Label(text='Buying Price:',
                              size_hint=(0.45, 0.25))
        self.bp_box.add_widget(self.bp_label)

        self.bp = ModifiedTextInput(hint_text='10000',
                                    size_hint=(0.55, 0.35),
                                    input_filter='int')
        self.bp_box.add_widget(self.bp)

        self.product_box = BoxLayout(orientation='vertical',
                                     size_hint=(0.9, 0.13),
                                     pos_hint={'x': 0.02, 'y': 0.59})

        self.product_label = Label(text='Product Name:',
                                   size_hint=(0.55, 0.25))
        self.product_box.add_widget(self.product_label)

        self.product_name = ModifiedTextInput(hint_text='Samsung S9+',
                                              size_hint=(1, 0.7))
        self.product_name.font_size = 15
        self.product_name.multiline = True
        self.product_box.add_widget(self.product_name)

        self.addbtn = AddButton(pos_hint={'x': 0.3, 'y': 0},
                                size_hint=(0.35, 0.14))
        self.add_widget(self.addbtn)
        self.addbtn.bind(on_press=self.callback)

    def callback(self, ins):
        print(ins)
        self.add_to_sheet()

    def add_to_sheet(self):
        self.background_color = get_color_from_hex('#9e2020')
        self.checked_box = ''

        if self.tglbtn1._get_active():
            self.checked_box = self.categories[0]

        elif self.tglbtn2._get_active():
            self.checked_box = self.categories[1]

        else:
            return self.error_popup(f'Please Choose {self.categories[0]} or {self.categories[1]}')

        if not self.price.text or not self.amt_input.text:
            return self.error_popup('Fields cannot be empty')

        if self.title.lower() == 'dollars':
            alltext = f'{self.checked_box.lower()} {self.title.lower()}'
            existing = AmountsField.existing_amounts_dict[alltext]
            company_take = AmountsField.existing_amounts_dict[f'{self.checked_box.lower()} dollars taken'][1]

            if (existing[0] - (int(self.amt_input.text) + company_take)) < 0:
                return self.error_popup()

            if (int(self.price.text)/int(self.amt_input.text)) < existing[1]:
                return self.error_popup('Cannot sell for less than buying price')

            existing[0] -= (float(self.amt_input.text) + company_take)
            AmountsField.labels_dict[alltext].text = f'{alltext.capitalize()}: {format(existing[0], ".2f")}'

            self.add_row(alltext.capitalize(), existing[1])

        elif self.title.lower() == 'days':
            ind = f'{self.checked_box.lower()} big cards'
            existing_bc = AmountsField.existing_amounts_dict[ind]
            if int(self.amt_input.text) > existing_bc[0]:
                return self.error_popup()

            days_price = AmountsField.existing_amounts_dict[f'{self.checked_box.lower()} {self.title.lower()}'][1]

            if (int(self.price.text)/int(self.amt_input.text)) < days_price:
                return self.error_popup('Cannot sell for less than buying price')

            existing_bc[0] -= int(self.amt_input.text)
            AmountsField.labels_dict[ind].text = f'{ind.capitalize()}: {int(existing_bc[0])}'
            
            self.add_row(f'{self.checked_box.lower()} {self.title.lower()}', days_price)

        elif self.title.lower() == 'cards':
            type_of_card = f'{self.checked_box.lower()} {self.card_choose.text.lower()} {self.title.lower()}'
            existing_card = AmountsField.existing_amounts_dict[type_of_card]

            if int(self.amt_input.text) > existing_card[0]:
                return self.error_popup()

            if (int(self.price.text)/int(self.amt_input.text)) < existing_card[1]:
                return self.error_popup('Cannot sell for less than buying price')

            existing_card[0] -= int(self.amt_input.text)
            AmountsField.labels_dict[type_of_card].text = f'{type_of_card.capitalize()}: {int(existing_card[0])}'
            self.add_row(type_of_card, existing_card[1])

        elif self.title.lower() == 'other':
            self.add_row(self.checked_box, int(self.bp.text), True)

        self.load_wb.save(self.wb_complete_path)
        with open('.storage.json', 'w') as f:
            json.dump(AmountsField.existing_amounts_dict, f)

    def error_popup(self, error_text='You can\'t sell more than what you have!'):
        popup_label = Label(text=error_text,
                            font_size=22)

        return Popup(title='Error',
                     content=popup_label,
                     size_hint=(0.45, 0.4)).open()

    def add_row(self, name, buying_price, is_phone_acc=False):
        max_row = self.ws.max_row + 1
        self.ws['A' + str(max_row)] = name
        self.ws['B' + str(max_row)] = f"{datetime.now().strftime('%H: %M: %S')}"
        self.ws['C' + str(max_row)] = self.amt_input.text
        self.ws['D' + str(max_row)] = self.price.text
        total_buying_price = (buying_price * int(self.amt_input.text))
        self.ws['E' + str(max_row)] = int(self.price.text) - total_buying_price
        self.ws['F' + str(max_row)] = total_buying_price
        TotalSelledField.totals[f"{datetime.now().strftime('%d')}"][1] += int(self.price.text)
        
        with open(TotalSelledField.totals_path, 'w') as f:
            json.dump(TotalSelledField.totals, f)

        if is_phone_acc:
            self.ws['G' + str(max_row)] = self.product_name.text
            self.product_name.text = ''
            self.bp.text = ''
        self.price.text = ''
        self.amt_input.text = '1'
        self.tglbtn1._set_active(False)
        self.tglbtn2._set_active(False)


class SpinnerOptions(Button):
    pass


class TotalSelledField(FloatLayout):
    now = datetime.now()
    totals_dir = f'storage/{now.strftime("%Y")}/{now.strftime("%m")}'
    totals_path = os.path.join(totals_dir, '.month_totals.json')
    if not Path(totals_path).is_file():
        with open(totals_path, 'w') as f:
            json.dump({}, f)
    
    with open(totals_path, 'r') as f:
        totals = json.load(f)
    
    tot_day = f"{now.strftime('%d')}"
    if tot_day not in totals.keys():
        with open(totals_path, 'w') as wf:
            totals[tot_day] = [0, 0]
            json.dump(totals, wf)
    tday = now.strftime('%d')
    mnth = now.strftime('%m')
    year = int(now.strftime('%Y'))

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.date = datetime.now().strftime('%a %d/%h/%Y')
        self.date_lbl = Label(text=f'[b][u]Date[/u]: {self.date}[/b]',
                              size_hint=(0.6, 0),
                              pos_hint={'x': 0, 'y': 0.9},
                              markup=True,
                              font_size=17)
        self.add_widget(self.date_lbl)

        self.time_lbl = Label(text=f'Time: {datetime.now().strftime("%H:%M")}',
                              size_hint=(0.25, 0),
                              pos_hint={'x': 0.7, 'y': 0.9},
                              font_size=19)
        self.add_widget(self.time_lbl)
        Clock.schedule_interval(self.time_update, 1)

        self.total_selled = Label(text=self.text_update(),
                                  size_hint=(0.5, 0),
                                  pos_hint={'x': 0.1, 'y': 0.5},
                                  font_size=17)
        self.add_widget(self.total_selled)
        self.show_btn = Button(text='[b]Show\nCharts[/b]',
                               pos_hint={'x': 0.6, 'y': 0},
                               size_hint=(0.23, 0.43),
                               background_normal='',
                               background_color=get_color_from_hex('#6E44FF'),
                               color=get_color_from_hex('#251333'),
                               font_size=18,
                               markup=True,)
        self.add_widget(self.show_btn)
        self.show_btn.bind(on_release=self.plot_it)

    def time_update(self, *args):
        self.time_lbl.text = f'Time: {datetime.now().strftime("%H:%M")}'
        self.total_selled.text = self.text_update()

    def text_update(self):
        return f"Total Selled: {self.totals[str(self.tday)][1]:,}"

    def plot_it(self, *args):
        self.fig, self.axs = plt.subplots(figsize=(15, 9))
        x_rng = [int(x) for x in self.totals.keys()]
        self.y_rng = self.totals.values()
        print(self.y_rng)
        plt.xticks(arange(1, max(x_rng) + 1, 1))
        plt.yticks(arange(0, 1000000, 200000))
        self.ordered_totals = [i[-1] for i in list(self.y_rng)]
        self.line = plt.plot(x_rng, self.ordered_totals, marker='o')
        saving_path = f"storage/{self.year}/{self.mnth}/{self.now.strftime('%B')}-graph.jpeg"

        self.annot = self.axs.annotate("", xy=(0, 0), xytext=(-20, 20), textcoords="offset points",
                                       bbox=dict(boxstyle="round", fc="w"),
                                       arrowprops=dict(arrowstyle="->"))
        self.annot.set_visible(False)
        
        self.fig.canvas.mpl_connect("motion_notify_event", self.hover)
        plt.show()

        self.axs.ticklabel_format(useOffset=False, style='plain')
        for i, j, w in zip(x_rng, self.ordered_totals, range(1, len(x_rng) + 1)):
            self.axs.annotate(str(j), xy=(i-0.1, j+0.1))

        self.fig.savefig(saving_path, bbox_inches='tight')
        

    def update_annot(self, ind):
        x, y = self.line[0].get_data()
        print(x, y)
        self.annot.xy = (x[ind["ind"][0]], y[ind["ind"][0]])
        text = f"Day {int(' '.join(list(map(str, ind['ind']))))+1}: {int(' '.join([str(self.ordered_totals[n]) for n in ind['ind']])):,}"
        print(text)
        self.annot.set_text(text)
        self.annot.get_bbox_patch().set_alpha(0.4)

    def hover(self, event):
        vis = self.annot.get_visible()
        if event.inaxes == self.axs:
            cont, ind = self.line[0].contains(event)
            if cont:
                self.update_annot(ind)
                self.annot.set_visible(True)
                self.fig.canvas.draw_idle()
            else:
                if vis:
                    self.annot.set_visible(False)
                    self.fig.canvas.draw_idle()
        

class FieldofFields(FloatLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.dol = EntryField('dollars',
                              'dollars',
                              size_hint=(0.33, 1))
        self.dol.pos_hint = {'x': 0, 'y': 0}
        self.add_widget(bg_change(pos_hint=self.dol.pos_hint,
                                  size_hint=self.dol.size_hint))
        self.add_widget(self.dol)

        self.days = EntryField('days',
                               'days',
                               size_hint=(0.33, 1))
        self.days.pos_hint = {'x': 0.335, 'y': 0}
        self.add_widget(bg_change(pos_hint=self.days.pos_hint,
                                  size_hint=self.days.size_hint))
        self.add_widget(self.days)

        self.cards = EntryField('cards',
                                'cards',
                                size_hint=(0.33, 1))
        self.cards.pos_hint = {'x': 0.67, 'y': 0}
        self.add_widget(bg_change(pos_hint=self.cards.pos_hint,
                                  size_hint=self.cards.size_hint))
        self.cards.amt_box.pos_hint = {'x': 0.02, 'y': 0.45}
        self.cards.price_box.pos_hint = {'x': 0.02, 'y': 0.3}
        self.cards.add_widget(self.cards.card_choose)

        self.add_widget(self.cards)


class EditPopup(Popup):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.size_hint = (0.45, 0.5)
        self.title = 'Edit'
        self.title_align = 'center'
        self.title_color = get_color_from_hex('#00FF00')
        self.title_size = 19

        self.inlayout = FloatLayout(size_hint=(1, 1),
                                    pos_hint={'x': 0, 'y': 0})

        self.labels = [txt.upper() for txt in AmountsField.existing_amounts_dict.keys()]

        self.to_be_edited = Spinner(size_hint=(0.45, 0.25),
                                    pos_hint={'x': 0.05, 'y': 0.6},
                                    values=self.labels,
                                    text=self.labels[0],
                                    sync_height=True)
        self.to_be_edited.bind(on_press=self.on_spinner_text)

        self.inlayout.add_widget(self.to_be_edited)
        self.add_widget(self.inlayout)

        # Plus Box.
        self.plus_box = BoxLayout(size_hint=(0.25, 0.1),
                                  orientation='horizontal',
                                  pos_hint={'x': 0.55, 'y': 0.77})

        self.plus_checkbox = CheckBox(size_hint=(0.2, 1),
                                      group='plusminus')
        self.plus_box.add_widget(self.plus_checkbox)

        self.plus_label = Label(text='Plus',
                                size_hint=(0.2, 1),
                                font_size=19)
        self.plus_box.add_widget(self.plus_label)
        self.inlayout.add_widget(self.plus_box)

        # Minus Box.
        self.minus_box = BoxLayout(size_hint=(0.25, 0.1),
                                   orientation='horizontal',
                                   pos_hint={'x': 0.55, 'y': 0.62})

        self.minus_checkbox = CheckBox(size_hint=(0.2, 1),
                                       group='plusminus')
        self.minus_box.add_widget(self.minus_checkbox)

        self.minus_label = Label(text='Minus',
                                 size_hint=(0.2, 1),
                                 font_size=19)
        self.minus_box.add_widget(self.minus_label)
        self.inlayout.add_widget(self.minus_box)

        # Price change.
        self.price_box = BoxLayout(size_hint=(0.35, 0.1),
                                   orientation='horizontal',
                                   pos_hint={'x': 0.555, 'y': 0.47})

        self.price_checkbox = CheckBox(size_hint=(0.2, 1),
                                       group='plusminus')
        self.price_box.add_widget(self.price_checkbox)

        self.price_label = Label(text='Price Change',
                                 size_hint=(0.4, 1),
                                 font_size=19)
        self.price_box.add_widget(self.price_label)
        self.inlayout.add_widget(self.price_box)
        
        self.input_label = Label(pos_hint={'x': 0.16, 'y': 0.33},
                                 size_hint=(0, 0),
                                 text='[b][u]Amount[/u]:[/b]',
                                 markup=True,
                                 font_size=20)
        self.inlayout.add_widget(self.input_label)

        self.input_num = ModifiedTextInput(input_filter='float',
                                           pos_hint={'x': 0.05, 'y': 0.15},
                                           size_hint=(0.5, 0.12))
        self.inlayout.add_widget(self.input_num)

        self.popup_btn = Button(pos_hint={'x': 0.7, 'y': 0.1},
                                size_hint=(0.25, 0.2),
                                text='Apply',
                                font_size=20)
        self.inlayout.add_widget(self.popup_btn)

        self.popup_btn.bind(on_press=self.on_popup_btn_press)
        self.price_checkbox.bind(on_press=self.price_cb_onpress)

    def on_spinner_text(self, spn):
        self.price_checkbox.state = 'normal'

    def price_cb_onpress(self, ins):
        self.input_num.text = str(AmountsField.existing_amounts_dict[self.to_be_edited.text.lower()][1])

    def on_popup_btn_press(self, btn):

        self.edited_title = self.to_be_edited.text.lower()
        self.current_value = AmountsField.existing_amounts_dict[self.edited_title]

        try:
            edit_number = float(self.input_num.text)
        except ValueError:
            edit_number = 1


        if self.edited_title in ['alfa days', 'touch days'] and self.price_checkbox.state != "down":
            days_add_text = '[b]Cannot Add/Remove Days![/b]'
            return self._label_popup('Error', days_add_text)
        
        elif self.minus_checkbox._get_active():
            self.current_value[0] -= edit_number
            self._update_text()

        elif self.plus_checkbox._get_active():
            self.current_value[0] += edit_number
            self._update_text()

        elif self.price_checkbox._get_active():
            self.current_value[1] = float(self.input_num.text)

        else:
            empty_field_text = 'Please Choose [b]Minus[/b], [b]Plus[/b]\nor [b]Change Price[/b]'
            return self._label_popup('Error', empty_field_text)

        with open('.storage.json', 'w') as f:
            json.dump(AmountsField.existing_amounts_dict, f)
            
    def _update_text(self):
        what_text = self.edited_title
        if what_text in ['alfa dollars', 'touch dollars']:
            text_to_be = f'{what_text.capitalize()}: {format(self.current_value[0], ".2f")}'
        else:
            text_to_be = f'{what_text.capitalize()}: {int(self.current_value[0])}'
        AmountsField.labels_dict[self.edited_title].text = text_to_be
        
    def _label_popup(self, title, text):
        return Popup(title=title,
                     size_hint=(0.35, 0.35),
                     content=Label(text=text,
                                   font_size=20,
                                   markup=True)).open()


class AmountsField(StackLayout):
    with open('.storage.json', 'r') as jf:
        existing_amounts_dict = json.load(jf)

    labels_dict = {}  # dictionary contains label objects of every sale type
    to_be_ignored = ['alfa days', 'touch days',
                     'alfa dollars taken', 'touch dollars taken']
    for i, v in existing_amounts_dict.items():
        if i in to_be_ignored:
            continue

        labels_dict[i] = Label(size_hint=(0.4, 0.2))
        if i in ['alfa dollars', 'touch dollars']:
            labels_dict[i].text = f'{i.capitalize()}: {format(v[0], ".2f")}'
        else:
            labels_dict[i].text = f'{i.capitalize()}: {int(v[0])}'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.orientation = 'tb-lr'

        for v in self.labels_dict.values():
            self.add_widget(v)

        self.edit_btn = Button(size_hint=(0.15, 1),
                               text='[b]EDIT[/b]',
                               font_size=20,
                               background_normal='',
                               background_color=get_color_from_hex('#6E44FF'),
                               color=get_color_from_hex('#231123'),
                               markup=True)
        self.add_widget(self.edit_btn)

        self.edit_btn.bind(on_press=self.popup_open)

    def popup_open(self, btn):
        edit_popup = EditPopup()
        edit_popup.open()


class MainScreen(FloatLayout):
    def __init__(self):
        super().__init__()

        self.amts = AmountsField(pos_hint={'x': 0.01, 'y': 0.75},
                                 size_hint=(0.6, 0.2))
        self.add_widget(bg_change(pos_hint=self.amts.pos_hint,
                                  size_hint=self.amts.size_hint,  # to remove later
                                  clr='#251333'))
        self.add_widget(self.amts)

        self.ddc = FieldofFields(size_hint=(0.65, 0.68),
                                 pos_hint={'x': 0.01, 'y': 0.015})
        self.add_widget(self.ddc)

        self.phone_acc = EntryField(title='other',
                                    grp='phoneacc',
                                    categories=['Phone', 'Accessory'],
                                    pos_hint={'x': 0.72, 'y': 0.015})
        self.phone_acc.size_hint = (0.22, 0.68)
        self.add_widget(bg_change(pos_hint=self.phone_acc.pos_hint,
                                  size_hint=self.phone_acc.size_hint))
        self.phone_acc.price_box.pos_hint = {'x': 0.02, 'y': 0.33}
        self.phone_acc.amt_box.pos_hint = {'x': 0.02, 'y': 0.46}
        self.phone_acc.add_widget(self.phone_acc.product_box)
        self.phone_acc.add_widget(self.phone_acc.bp_box)
        self.add_widget(self.phone_acc)

        self.totals_field = TotalSelledField(size_hint=(0.35, 0.2),
                                             pos_hint={'x': 0.63, 'y': 0.75})
        self.add_widget(bg_change(pos_hint=self.totals_field.pos_hint,
                                  size_hint=self.totals_field.size_hint,  # to remove later
                                  clr='#251333'))
        self.add_widget(self.totals_field)


class ManagerApp(App):
    def build(self):
        Window.bind(on_request_close=self.on_request_close)

        self.main_screen = MainScreen()
        self.main_screen.size_hint_max = (1050, 750)
        self.main_screen.size_hint_min = (700, 600)
        Window.clearcolor = get_color_from_hex('#231123')  # 230C33 101D42 DCF763
        Window.size = (950, 720)
        return self.main_screen
    
    def on_request_close(self, *args):
        self.textpopup('Exit', 'Are You Sure?')
        return True

    def textpopup(self, title='', text=''):

        box = FloatLayout()
        box.add_widget(Label(text=text,
                             size_hint=(1, 0),
                             pos_hint={'x': 0, 'y': 0.6},
                             font_size=26))
        exit_button = Button(text='Yes',
                             size_hint=(0.5, 0.25),
                             pos_hint={'x': 0, 'y': 0},
                             font_size=20)
        box.add_widget(exit_button)
        popup = Popup(title=title,
                      content=box,
                      size_hint=(0.5, 0.5))
        exit_button.bind(on_release=self.stop)
        
        cancel_button = Button(text='No',
                               size_hint=(0.5, 0.25),
                               pos_hint={'x': 0.5, 'y': 0},
                               font_size=20)
        box.add_widget(cancel_button)
        cancel_button.bind(on_release=popup.dismiss)
        popup.open()


if __name__ == '__main__':
    myapp = ManagerApp()
    myapp.run()
